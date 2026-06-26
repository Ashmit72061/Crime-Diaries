import json
import os
import pandas as pd
import openpyxl
from datetime import datetime
from db import engine
from sqlalchemy import text

TEMPLATES_DIR = os.path.join(os.path.dirname(__file__), 'templates')

def load_local_template(template_id):
    """Load a template definition from python_worker/templates/<id>.json if it exists."""
    path = os.path.join(TEMPLATES_DIR, f'{template_id}.json')
    if os.path.exists(path):
        with open(path, encoding='utf-8') as f:
            return json.load(f)
    return None

def load_template(template_id, engine):
    with engine.connect() as conn:
        row = conn.execute(
            text("SELECT id, name_en, template_type, template_definition, applicable_record_types FROM report_templates WHERE id = :id"),
            {'id': template_id}
        ).fetchone()

    if not row:
        return None

    raw_def = json.loads(row[3]) if isinstance(row[3], str) else (row[3] or {})
    applicable = json.loads(row[4]) if isinstance(row[4], str) else (row[4] or ['CASE'])
    record_type = applicable[0] if applicable else 'CASE'

    # DB templates store definition as {layout, header, sections[{fields}]}.
    # Python expects {filter_spec, fixed_fields, header}.  Normalize here.
    if 'filter_spec' not in raw_def:
        fixed_fields = []
        for section in raw_def.get('sections', []):
            fixed_fields.extend(section.get('fields', []))
        raw_def = {
            'filter_spec': {'record_type': record_type},
            'fixed_fields': fixed_fields,
            'header': raw_def.get('header', {'title_en': row[1]}),
            'template_type': row[2] or 'PROFORMA',
        }

    return {
        'id': row[0],
        'name_en': row[1],
        'template_type': row[2] or 'PROFORMA',
        'template_definition': raw_def,
    }

def get_predefined_definition(template_id):
    local = load_local_template(template_id)
    if local:
        return local
    if template_id == 'arrest-summary':
        return {
            'filter_spec': { 'record_type': 'ARREST' },
            'fixed_fields': ['uid', 'arrest_date', 'arrested_name', 'crime_head', 'io_name'],
            'header': { 'title_en': 'Arrest Summary Report', 'short_name_en': 'Arrest Summary' }
        }
    elif template_id == 'pcr-call-log':
        return {
            'filter_spec': { 'record_type': 'PCR_CALL' },
            'fixed_fields': ['pcr_no', 'gd_date', 'caller_mobile', 'occurrence_place', 'call_head', 'status'],
            'header': { 'title_en': 'PCR Call Log', 'short_name_en': 'PCR Call Log' }
        }
    elif template_id == 'cases-register':
        return {
            'filter_spec': { 'record_type': 'CASE' },
            'fixed_fields': ['fir_no', 'fir_date', 'complainant_name', 'local_head', 'brief_facts'],
            'header': { 'title_en': 'Cases Register', 'short_name_en': 'Cases Register' }
        }
    else:
        return {
            'filter_spec': { 'record_type': 'CASE' },
            'fixed_fields': ['uid', 'fir_no', 'fir_date'],
            'header': { 'title_en': 'PHAROS REPORT', 'short_name_en': 'Report' }
        }

def load_field_registry(field_keys, engine):
    if not field_keys:
        return {}
    with engine.connect() as conn:
        res = conn.execute(
            text("SELECT field_key, label_en, label_hi, field_type FROM field_registry WHERE field_key IN :keys"),
            {'keys': tuple(field_keys)}
        ).fetchall()
        
    meta = {}
    for r in res:
        meta[r[0]] = {
            'label_en': r[1],
            'label_hi': r[2],
            'field_type': r[3]
        }
    return meta

def query_records(definition, user_filters, engine):
    record_type = definition['filter_spec']['record_type']
    data_filter = definition['filter_spec'].get('data_filter', {})

    field_keys = definition.get('fixed_fields', [])
    selects = []

    for k in field_keys:
        selects.append(f"(records.data::jsonb)->>'{k}' as {k}")
            
    # Include record date and ps name in raw queries
    select_expr = ", ".join(selects) if selects else "records.id"
    sql = f"""
        SELECT {select_expr}, hn.name_en as ps_name, records.record_date
        FROM records
        LEFT JOIN hierarchy_nodes hn ON records.ps_id = hn.id
        WHERE records.record_type = :record_type
    """
    
    params = {
        'record_type': record_type,
        'date_from': user_filters.get('date_from', user_filters.get('from_date', '2020-01-01')),
        'date_to': user_filters.get('date_to', user_filters.get('to_date', '2030-01-01'))
    }
    
    # Date filters
    sql += " AND records.record_date BETWEEN :date_from AND :date_to"
    
    # PS filter
    if user_filters.get('ps_id'):
        sql += " AND records.ps_id = :ps_id"
        params['ps_id'] = user_filters['ps_id']
    elif user_filters.get('psId'):
        sql += " AND records.ps_id = :psId"
        params['psId'] = user_filters['psId']
        
    # Apply data filters
    if data_filter:
        for k, v in data_filter.items():
            sql += f" AND records.data::jsonb @> :{k}_val::jsonb"
            params[f"{k}_val"] = json.dumps({k: v})

    # Apply dynamic user filters
    system_keys = {'date_from', 'from_date', 'date_to', 'to_date', 'ps_id', 'psId', 'district_id', 'districtId', 'selected_sub_templates', 'page', 'limit'}
    core_columns = {'id', 'current_status', 'current_level'}

    for k, v in user_filters.items():
        if k in system_keys or v is None or v == '':
            continue

        if k in core_columns:
            sql += f" AND records.{k} = :{k}_user_val"
            params[f"{k}_user_val"] = v
        else:
            sql += f" AND records.data::jsonb @> :{k}_user_val::jsonb"
            params[f"{k}_user_val"] = json.dumps({k: v})

    with engine.connect() as conn:
        df = pd.read_sql(text(sql), conn, params=params)
        
    # Load field registry metadata for header renames
    fields_meta = load_field_registry(field_keys, engine)
    
    # Add Sr. No. column at the start
    if not df.empty:
        df.insert(0, 'Sr. No.', range(1, len(df) + 1))
        
        # Rename columns to label_en from registry
        rename_map = {k: fields_meta.get(k, {}).get('label_en', k.replace('_', ' ').title()) for k in field_keys}
        df.rename(columns=rename_map, inplace=True)
        
        # Drop raw fields like ps_name and record_date if they are not explicitly asked in fixed_fields
        drop_cols = []
        if 'ps_name' not in field_keys:
            drop_cols.append('ps_name')
        if 'record_date' not in field_keys:
            drop_cols.append('record_date')
        df.drop(columns=[c for c in drop_cols if c in df.columns], inplace=True)
        
    return df

def write_header_rows(ws, header_def, user_filters, engine):
    title_en = header_def.get('title_en', 'PHAROS REPORT')
    title_hi = header_def.get('title_hi', '')
    
    ws.append([f"{title_en} {title_hi}".strip()])
    ws.append([f"Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"])
    
    ps_id = user_filters.get('ps_id') or user_filters.get('psId')
    jurisdiction = "All jurisdictions"
    if ps_id:
        with engine.connect() as conn:
            r = conn.execute(text("SELECT name_en FROM hierarchy_nodes WHERE id = :id"), {'id': ps_id}).fetchone()
        if r:
            jurisdiction = r[0]
                
    date_from = user_filters.get('date_from', user_filters.get('from_date', 'N/A'))
    date_to = user_filters.get('date_to', user_filters.get('to_date', 'N/A'))
    date_range = f"Date Range: {date_from} to {date_to}"
    ws.append([f"Jurisdiction: {jurisdiction} | {date_range}"])
    ws.append([]) # empty separator row
    
    from openpyxl.styles import Font
    ws.cell(row=1, column=1).font = Font(name='Arial', size=13, bold=True, color='1F4E79')
    ws.cell(row=2, column=1).font = Font(name='Arial', size=9, italic=True, color='595959')
    ws.cell(row=3, column=1).font = Font(name='Arial', size=9, color='595959')

def write_dataframe(ws, df, start_row=5, engine=None):
    headers = list(df.columns)
    ws.append(headers)
    
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9')
    )
    
    header_row_idx = ws.max_row
    ws.row_dimensions[header_row_idx].height = 25
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=header_row_idx, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_align
        cell.border = thin_border
        
    data_font = Font(name='Arial', size=9)
    data_align_left = Alignment(horizontal='left', vertical='center')
    data_align_center = Alignment(horizontal='center', vertical='center')
    zebra_fill = PatternFill(start_color='F2F6F9', end_color='F2F6F9', fill_type='solid')
    
    for row_idx, row in enumerate(df.itertuples(index=False), start=header_row_idx + 1):
        ws.row_dimensions[row_idx].height = 20
        use_zebra = (row_idx % 2 == 0)
        
        for col_idx, val in enumerate(row, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin_border
            if use_zebra:
                cell.fill = zebra_fill
            
            if isinstance(val, (int, float)) or str(val).startswith('Sr. No.'):
                cell.alignment = data_align_center
            else:
                cell.alignment = data_align_left
                
    for col in ws.columns:
        max_len = 10
        for cell in col:
            if cell.row < header_row_idx:
                continue
            val_str = str(cell.value or '')
            if len(val_str) > max_len:
                max_len = len(val_str)
        col_letter = openpyxl.utils.get_column_letter(col[0].column)
        ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

def generate_single_sheet_workbook(definition, user_filters, engine):
    df = query_records(definition, user_filters, engine)
    
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = definition.get('header', {}).get('short_name_en', 'Report')[:30]
    
    write_header_rows(ws, definition.get('header', {}), user_filters, engine)
    
    if df.empty:
        ws.append(["No records found for the selected period and filters."])
        ws.append([f"Filter applied: {json.dumps(definition.get('filter_spec', {}))}"])
    else:
        write_dataframe(ws, df, start_row=5, engine=engine)
        
    return wb

def generate_composite(definition, user_filters, engine):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    sub_ids = definition.get('sub_template_ids', [])
    selected_sub_templates = user_filters.get('selected_sub_templates')
    
    if selected_sub_templates:
        sub_ids = [sid for sid in sub_ids if sid in selected_sub_templates]
        
    for sub_id in sub_ids:
        template = load_template(sub_id, engine)
        if not template:
            # Check if it maps to predefined ones as fallback
            sub_def = get_predefined_definition(sub_id)
            sub_name = sub_id.replace('-', ' ').title()
        else:
            sub_def = template['template_definition']
            sub_name = template['name_en']
            
        sheet_name = sub_def.get('header', {}).get('short_name_en', sub_name)
        for char in ':\\/?*[]':
            sheet_name = sheet_name.replace(char, '')
        sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        df = query_records(sub_def, user_filters, engine)
        
        write_header_rows(ws, sub_def.get('header', {}), user_filters, engine)
        
        if df.empty:
            ws.append(["No records found for the selected period and filters."])
            ws.append([f"Filter applied: {json.dumps(sub_def.get('filter_spec', {}))}"])
        else:
            write_dataframe(ws, df, start_row=5, engine=engine)
            
    return wb

def generate_custom_workbook(custom_definition, user_filters, engine):
    wb = openpyxl.Workbook()
    wb.remove(wb.active)
    
    sheets = custom_definition.get('sheets', [])
    if not sheets:
        wb.create_sheet(title="Report").append(["No sheets defined in custom report definition."])
        return wb
        
    for i, sheet in enumerate(sheets):
        record_type = sheet.get('record_type')
        field_keys = sheet.get('field_keys', [])
        sheet_title = sheet.get('sheet_name', f"Sheet {i+1}")
        
        sheet_def = {
            'filter_spec': { 'record_type': record_type },
            'fixed_fields': field_keys,
            'header': {
                'title_en': custom_definition.get('title_en', 'Custom Report'),
                'short_name_en': sheet_title
            }
        }
        
        sheet_name = sheet_title
        for char in ':\\/?*[]':
            sheet_name = sheet_name.replace(char, '')
        sheet_name = sheet_name[:31]
        
        ws = wb.create_sheet(title=sheet_name)
        df = query_records(sheet_def, user_filters, engine)
        
        write_header_rows(ws, sheet_def.get('header', {}), user_filters, engine)
        
        if df.empty:
            ws.append(["No records found for the selected period and filters."])
            ws.append([f"Filter applied: {json.dumps(sheet_def.get('filter_spec', {}))}"])
        else:
            write_dataframe(ws, df, start_row=5, engine=engine)
            
    return wb

def query_linked_records(definition, user_filters, engine):
    direction = definition.get('direction', 'source_to_target')
    link_code = definition.get('link_type_code', 'CASE_ARREST')

    def jf(col, key):
        return f"({col}::jsonb)->>'{key}'"

    params = {
        'date_from': user_filters.get('date_from', user_filters.get('from_date', '2020-01-01')),
        'date_to':   user_filters.get('date_to',   user_filters.get('to_date',   '2030-01-01')),
        'link_code': link_code,
    }

    if direction == 'target_to_source':
        # ARREST primary, LEFT JOIN to parent CASE
        coalesce = f"COALESCE({jf('c.data', 'fir_no')}, {jf('a.data', 'linked_fir_dd_no')})"
        sql = f"""
            SELECT
              {jf('a.data', 'arrested_name')}    AS arrested_name,
              {jf('a.data', 'parents_name')}     AS parents_name,
              {jf('a.data', 'arrested_address')} AS arrested_address,
              {jf('a.data', 'age_gender')}       AS age_gender,
              {coalesce}                          AS fir_dd_no,
              {jf('a.data', 'arrest_date')}      AS arrest_date,
              {jf('a.data', 'sections')}         AS sections,
              hn.name_en                          AS ps_name,
              {jf('a.data', 'io_name')}          AS io_name,
              {jf('a.data', 'io_rank')}          AS io_rank,
              {jf('a.data', 'io_mobile')}        AS io_mobile,
              {jf('a.data', 'status')}           AS status
            FROM records a
            JOIN hierarchy_nodes hn ON hn.id = a.ps_id
            LEFT JOIN record_links rl
              ON rl.target_record_id = a.id
            LEFT JOIN link_type_registry ltr
              ON ltr.id = rl.link_type_id AND ltr.code = :link_code
            LEFT JOIN records c ON c.id = rl.source_record_id
            WHERE a.record_type = 'ARREST'
              AND a.record_date BETWEEN :date_from AND :date_to
        """
        if user_filters.get('ps_id') or user_filters.get('psId'):
            sql += " AND a.ps_id = :ps_id"
            params['ps_id'] = user_filters.get('ps_id') or user_filters.get('psId')
        elif user_filters.get('district_id') or user_filters.get('districtId'):
            sql += " AND a.district_id = :district_id"
            params['district_id'] = user_filters.get('district_id') or user_filters.get('districtId')
        sql += " ORDER BY a.record_date DESC"

    else:
        # CASE primary, LEFT JOIN to linked ARRESTed persons
        sql = f"""
            SELECT
              hn.name_en                                   AS ps_name,
              {jf('c.data', 'fir_no')}                    AS fir_no,
              {jf('c.data', 'sections')}                  AS sections,
              {jf('c.data', 'complainant_name')}          AS complainant_name,
              {jf('c.data', 'complainant_parent_name')}   AS complainant_parent_name,
              {jf('c.data', 'complainant_address')}       AS complainant_address,
              {jf('c.data', 'time_of_occurrence')}        AS time_of_occurrence,
              {jf('c.data', 'occurrence_place')}          AS occurrence_place,
              {jf('c.data', 'occurrence_date')}           AS occurrence_date,
              {jf('c.data', 'brief_facts')}               AS brief_facts,
              {jf('a.data', 'arrested_name')}             AS arrested_name,
              {jf('a.data', 'parents_name')}              AS parents_name,
              {jf('a.data', 'arrested_address')}          AS arrested_address,
              {jf('a.data', 'age_gender')}                AS age_gender,
              {jf('c.data', 'io_name')}                   AS io_name
            FROM records c
            JOIN hierarchy_nodes hn ON hn.id = c.ps_id
            LEFT JOIN record_links rl
              ON rl.source_record_id = c.id
            LEFT JOIN link_type_registry ltr
              ON ltr.id = rl.link_type_id AND ltr.code = :link_code
            LEFT JOIN records a ON a.id = rl.target_record_id
            WHERE c.record_type = 'CASE'
              AND c.record_date BETWEEN :date_from AND :date_to
        """
        if user_filters.get('ps_id') or user_filters.get('psId'):
            sql += " AND c.ps_id = :ps_id"
            params['ps_id'] = user_filters.get('ps_id') or user_filters.get('psId')
        elif user_filters.get('district_id') or user_filters.get('districtId'):
            sql += " AND c.district_id = :district_id"
            params['district_id'] = user_filters.get('district_id') or user_filters.get('districtId')
        sql += " ORDER BY c.record_date DESC"

    with engine.connect() as conn:
        df = pd.read_sql(text(sql), conn, params=params)

    # Map column order to match definition['columns']
    col_keys = [c['key'] for c in definition.get('columns', [])]
    ordered_cols = [k for k in col_keys if k in df.columns]
    df = df[ordered_cols] if ordered_cols else df

    if not df.empty:
        df.insert(0, 'Sr. No.', range(1, len(df) + 1))

    # Rename to display labels
    label_map = {c['key']: c['label'] for c in definition.get('columns', [])}
    df.rename(columns=label_map, inplace=True)

    return df


def generate_linked_workbook(definition, user_filters, engine):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    df = query_linked_records(definition, user_filters, engine)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = definition.get('header', {}).get('short_name_en', 'Report')[:30]

    write_header_rows(ws, definition.get('header', {}), user_filters, engine)

    header_groups = definition.get('header_groups', [])

    dark_blue  = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    mid_blue   = PatternFill(start_color='2E75B6', end_color='2E75B6', fill_type='solid')
    white_bold = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    white_norm = Font(name='Arial', size=9, bold=False, color='FFFFFF')
    center     = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),  bottom=Side(style='thin', color='D9D9D9')
    )

    # Row 4: group headers with merged cells
    group_row = ws.max_row + 1
    ws.row_dimensions[group_row].height = 25
    col = 1
    for group in header_groups:
        span = group.get('span', 1)
        end = col + span - 1
        if span > 1:
            ws.merge_cells(start_row=group_row, start_column=col, end_row=group_row, end_column=end)
        cell = ws.cell(row=group_row, column=col, value=group['label'])
        cell.fill = dark_blue
        cell.font = white_bold
        cell.alignment = center
        cell.border = thin_border
        col = end + 1

    # Row 5: sub-column labels (one per actual data column, Sr. No. first)
    sub_row = group_row + 1
    ws.row_dimensions[sub_row].height = 22
    sub_labels = ['Sr. No.'] + [c['label'] for c in definition.get('columns', [])]
    for ci, label in enumerate(sub_labels, start=1):
        cell = ws.cell(row=sub_row, column=ci, value=label)
        cell.fill = mid_blue
        cell.font = white_norm
        cell.alignment = center
        cell.border = thin_border

    if df.empty:
        ws.append(['No records found for the selected period and filters.'])
    else:
        # write data rows manually (write_dataframe starts from current max_row+1)
        data_font  = Font(name='Arial', size=9)
        left_align = Alignment(horizontal='left',   vertical='center')
        ctr_align  = Alignment(horizontal='center', vertical='center')
        zebra_fill = PatternFill(start_color='F2F6F9', end_color='F2F6F9', fill_type='solid')

        for row_idx, row in enumerate(df.itertuples(index=False), start=sub_row + 1):
            ws.row_dimensions[row_idx].height = 20
            use_zebra = (row_idx % 2 == 0)
            for ci, val in enumerate(row, start=1):
                cell = ws.cell(row=row_idx, column=ci, value=val)
                cell.font = data_font
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill
                cell.alignment = ctr_align if isinstance(val, (int, float)) else left_align

        # Auto-fit column widths
        for col_cells in ws.columns:
            max_len = 10
            for cell in col_cells:
                if cell.row < group_row:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    return wb


def render_pdf(df, definition, filters, file_path):
    title = definition.get('header', {}).get('title_en', 'PHAROS REPORT')
    html_content = f"""<!DOCTYPE html>
<html>
<head>
    <meta charset="utf-8">
    <style>
        body {{ font-family: Arial, sans-serif; padding: 20px; color: #333; }}
        h1 {{ color: #1F4E79; font-size: 18px; margin-bottom: 5px; }}
        p {{ color: #595959; font-size: 10px; margin: 2px 0; }}
        table {{ width: 100%; border-collapse: collapse; margin-top: 15px; }}
        th, td {{ border: 1px solid #D9D9D9; padding: 6px 8px; text-align: left; font-size: 9px; }}
        th {{ background-color: #1F4E79; color: white; font-weight: bold; }}
        tr:nth-child(even) {{ background-color: #F2F6F9; }}
    </style>
</head>
<body>
    <h1>{title}</h1>
    <p>Generated At: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}</p>
    <table>
        <thead>
            <tr>
                {"".join(f"<th>{col}</th>" for col in df.columns)}
            </tr>
        </thead>
        <tbody>
            {"".join(
                f"<tr>{''.join(f'<td>{str(val or "")}</td>' for val in row)}</tr>"
                for row in df.itertuples(index=False)
            )}
        </tbody>
    </table>
</body>
</html>"""
    try:
        from weasyprint import HTML
        HTML(string=html_content).write_pdf(file_path)
        print("[Worker] PDF generated successfully using WeasyPrint.")
    except Exception as e:
        print(f"[Worker] WeasyPrint generation failed ({e}). Saving HTML layout directly to disk.")
        with open(file_path, 'w', encoding='utf-8') as f:
            f.write(html_content)

DAILY_DIARY_PARALLEL_TEMPLATE_IDS = {
    'daily-diary', 'dd-manual-fir', 'dd-eburglary-ehouse-theft-mvt',
    'dd-arrested-all-heads', 'dd-arrested-east-district', 'dd-arrested-kalandara',
    'dd-arrested-efir-theft', 'dd-arrested-efir-mv-theft', 'dd-proclaimed-offenders',
    'dd-arrested-24hrs', 'dd-missing-uidb', 'dd-women-children-missing',
    'dd-preventive-action', 'dd-inquest-registered', 'dd-important-cases',
    'dd-fir-goswara-summary', 'dd-financial-fraud-arrest', 'dd-ndps-action',
}

TEMPLATE_TO_TABLE_NAMES = {
    'daily-diary': None,
    'dd-manual-fir': ['excel_1manual_fir'],
    'dd-eburglary-ehouse-theft-mvt': ['excel_2eburglary_cases', 'excel_3ehouse_theft_cases', 'excel_4eother_theft_cases', 'excel_5mvt_cases'],
    'dd-arrested-all-heads': ['excel_6arrested_all_heads'],
    'dd-arrested-east-district': ['excel_7arrested_east_district'],
    'dd-arrested-kalandara': ['excel_8arrested_kalandara'],
    'dd-arrested-efir-theft': ['excel_9arrested_efir_theft'],
    'dd-arrested-efir-mv-theft': ['excel_10arrested_efir_mv_theft'],
    'dd-proclaimed-offenders': ['excel_11proclaimed_offenders'],
    'dd-arrested-24hrs': ['excel_13arrested_24_hrs_list'],
    'dd-missing-uidb': ['excel_18missing_persons', 'excel_19uidb', 'excel_20abandoned_persons', 'excel_21traced_persons'],
    'dd-women-children-missing': ['excel_22women_missing', 'excel_23children_missing'],
    'dd-preventive-action': ['excel_24preventive_action'],
    'dd-inquest-registered': ['excel_25inquest_registered', 'excel_26inquest_acpsdm_disposal'],
    'dd-important-cases': ['excel_27important_cases'],
    'dd-fir-goswara-summary': ['excel_28fir_goswara_summary'],
    'dd-financial-fraud-arrest': ['excel_29financial_fraud_arrest'],
    'dd-ndps-action': ['excel_31ndps_action'],
}

def generate_daily_diary_workbook(custom_definition):
    """
    Build the daily diary Excel workbook from pre-classified data.
    custom_definition must contain: type='DAILY_DIARY', date, reports, report_columns, sheets.
    Node.js classifies; Python formats — no DB queries needed here.
    """
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    reports = custom_definition.get('reports', [])
    report_columns = custom_definition.get('report_columns', {})
    sheets_data = custom_definition.get('sheets', {})
    date = custom_definition.get('date', '')

    header_fill = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
    header_font = Font(name='Arial', size=9, bold=True, color='FFFFFF')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    data_font = Font(name='Arial', size=9)
    data_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    zebra_fill = PatternFill(start_color='EBF3FA', end_color='EBF3FA', fill_type='solid')
    thin_border = Border(
        left=Side(style='thin', color='D9D9D9'),
        right=Side(style='thin', color='D9D9D9'),
        top=Side(style='thin', color='D9D9D9'),
        bottom=Side(style='thin', color='D9D9D9'),
    )

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for rep in reports:
        table_name = rep['tableName']
        label = rep['label']
        num = rep['num']

        # Sheet title: "1. Manual FIR"
        sheet_title = f"{num}. {label}"
        for bad_char in ':\\/?*[]':
            sheet_title = sheet_title.replace(bad_char, '')
        sheet_title = sheet_title[:31]

        ws = wb.create_sheet(title=sheet_title)

        # Title row
        ws.append([f"PHAROS Daily Diary — {label}"])
        ws.append([f"Date: {date}"])
        ws.append([])  # spacer

        col_keys = report_columns.get(table_name, [])
        col_headers = [k.replace('_', ' ').title() for k in col_keys]

        # Header row
        ws.append(col_headers if col_headers else ['No Data'])
        header_row_idx = ws.max_row
        ws.row_dimensions[header_row_idx].height = 22
        for ci in range(1, len(col_headers) + 1):
            cell = ws.cell(row=header_row_idx, column=ci)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_align
            cell.border = thin_border

        # Title style
        ws.cell(row=1, column=1).font = Font(name='Arial', size=11, bold=True, color='1F3864')
        ws.cell(row=2, column=1).font = Font(name='Arial', size=9, italic=True, color='595959')

        # Data rows
        rows = sheets_data.get(table_name, [])
        for row_idx_offset, row_obj in enumerate(rows, start=1):
            row_values = [row_obj.get(k, '') for k in col_keys]
            ws.append(row_values)
            actual_row = header_row_idx + row_idx_offset
            ws.row_dimensions[actual_row].height = 18
            use_zebra = (row_idx_offset % 2 == 0)
            for ci, val in enumerate(row_values, start=1):
                cell = ws.cell(row=actual_row, column=ci)
                cell.font = data_font
                cell.alignment = data_align
                cell.border = thin_border
                if use_zebra:
                    cell.fill = zebra_fill

        if not rows:
            ws.append(['No records for this date.'])

        # Auto-fit column widths (capped at 45)
        for col_cells in ws.columns:
            max_len = 10
            for cell in col_cells:
                if cell.row < header_row_idx:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    return wb


def generate_report(job_id):
    print(f"[Worker] Starting report job: {job_id}")
    with engine.connect() as conn:
        job = conn.execute(
            text("SELECT template_id, custom_definition, filters, format, file_path, created_by FROM report_jobs WHERE id = :id"),
            {'id': job_id}
        ).fetchone()
        
    if not job:
        raise Exception(f"Report job {job_id} not found in database")
        
    template_id = job[0]
    custom_definition = job[1]
    filters = job[2]
    format_type = job[3].upper()
    file_path = job[4]
    user_id = job[5]
    
    if not os.path.isabs(file_path):
        backend_dir = os.path.abspath(os.path.join(os.path.dirname(__file__), '../backend'))
        file_path = os.path.abspath(os.path.join(backend_dir, file_path))
        
    if isinstance(custom_definition, str):
        custom_definition = json.loads(custom_definition)
    if isinstance(filters, str):
        filters = json.loads(filters)
        
    template = None
    if template_id:
        template = load_template(template_id, engine)
        if not template or not template.get('template_definition') or 'filter_spec' not in template['template_definition']:
            print(f"[Worker] Predefined template row {template_id} not found or incomplete in DB. Falling back to predefined structures.")
            template = {
                'id': template_id,
                'name_en': template_id.replace('-', ' ').title(),
                'template_type': 'PROFORMA',
                'template_definition': get_predefined_definition(template_id)
            }
        definition = template['template_definition']
    else:
        definition = custom_definition
        # Map first sheet for CSV/PDF formats
        if format_type in ['CSV', 'PDF']:
            sheets = custom_definition.get('sheets', []) if custom_definition else []
            if sheets:
                first_sheet = sheets[0]
                definition = {
                    'filter_spec': { 'record_type': first_sheet.get('record_type') },
                    'fixed_fields': first_sheet.get('field_keys', []),
                    'header': {
                        'title_en': custom_definition.get('title_en', 'Custom Report'),
                        'short_name_en': first_sheet.get('sheet_name', 'Report')
                    }
                }
        
    dir_name = os.path.dirname(file_path)
    if dir_name and not os.path.exists(dir_name):
        os.makedirs(dir_name, exist_ok=True)
        
    template_type = definition.get('template_type', 'PROFORMA') if definition else 'PROFORMA'

    if format_type == 'CSV':
        if template_type in ('COMPOSITE', 'LINKED'):
            raise Exception(f"CSV format not supported for {template_type} templates")
        df = query_records(definition, filters, engine)
        df.to_csv(file_path, index=False)

    elif format_type in ['EXCEL', 'XLSX']:
        if not template and custom_definition and custom_definition.get('type') == 'DAILY_DIARY':
            wb = generate_daily_diary_workbook(custom_definition)
            wb.save(file_path)
        elif template_id in DAILY_DIARY_PARALLEL_TEMPLATE_IDS:
            import subprocess
            table_names = TEMPLATE_TO_TABLE_NAMES.get(template_id)
            cli_filters = dict(filters) if filters else {}
            if table_names:
                cli_filters['tableNames'] = table_names
            script_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '../backend/scripts/dev/generate_parallel_report_cli.js'))
            filters_json = json.dumps(cli_filters)
            cmd = ['node', script_path, job_id, filters_json, file_path]
            print(f"[Worker] Delegating to Node.js parallel engine for template: {template_id}")
            subprocess.run(cmd, check=True)
        elif template_type == 'LINKED':
            wb = generate_linked_workbook(definition, filters, engine)
            wb.save(file_path)
        elif template_type == 'COMPOSITE' or (template and template.get('template_type') == 'COMPOSITE'):
            wb = generate_composite(definition, filters, engine)
            wb.save(file_path)
        elif not template:
            wb = generate_custom_workbook(definition, filters, engine)
            wb.save(file_path)
        else:
            wb = generate_single_sheet_workbook(definition, filters, engine)
            wb.save(file_path)

    elif format_type == 'PDF':
        if template_type == 'LINKED':
            df = query_linked_records(definition, filters, engine)
        else:
            df = query_records(definition, filters, engine)
        render_pdf(df, definition, filters, file_path)
        
    with engine.begin() as conn:
        conn.execute(
            text("UPDATE report_jobs SET status='READY', updated_at=:now WHERE id=:id"),
            {'now': datetime.now().isoformat(), 'id': job_id}
        )

    file_size = os.path.getsize(file_path)
    if os.getenv('PHAROS_TEST') == 'true':
        print(f"[Worker] Skipping event publishing in test mode for job: {job_id}")
    else:
        try:
            from events import publish_event
            publish_event('report.generated', {
                'job_id': job_id,
                'template_id': template_id,
                'requested_by': user_id,
                'file_path': file_path,
                'format': format_type,
                'file_size_bytes': file_size
            })
        except Exception as err:
            print(f"[WorkerError] Failed to publish event: {err}")
    print(f"[Worker] Completed report job: {job_id}. File size: {file_size} bytes")
