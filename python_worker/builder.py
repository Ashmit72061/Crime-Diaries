import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

_HEADER_FILL  = PatternFill(start_color='1F3864', end_color='1F3864', fill_type='solid')
_HEADER_FONT  = Font(name='Arial', size=9, bold=True, color='FFFFFF')
_HEADER_ALIGN = Alignment(horizontal='center', vertical='center', wrap_text=True)
_DATA_FONT    = Font(name='Arial', size=9)
_DATA_ALIGN   = Alignment(horizontal='left', vertical='center', wrap_text=True)
_ZEBRA_FILL   = PatternFill(start_color='EBF3FA', end_color='EBF3FA', fill_type='solid')
_THIN_BORDER  = Border(
    left=Side(style='thin', color='D9D9D9'),
    right=Side(style='thin', color='D9D9D9'),
    top=Side(style='thin', color='D9D9D9'),
    bottom=Side(style='thin', color='D9D9D9'),
)
_TITLE_FONT   = Font(name='Arial', size=11, bold=True, color='1F3864')
_META_FONT    = Font(name='Arial', size=9, italic=True, color='595959')


def _resolve_header(col_key, sheet_column_labels, shared_labels):
    """Look up display label: sheet overrides > shared labels > title-case fallback."""
    if col_key in sheet_column_labels:
        return sheet_column_labels[col_key]
    if col_key in shared_labels:
        return shared_labels[col_key]
    return col_key.replace('_', ' ').title()


SHARED_COLUMN_LABELS = {
    'sn': 'S.N.', 'sno': 'S.No.', 's_no': 'S. No.', 'sr_no': 'Sr. No.', 'sr': 'Sr.',
    'ps': 'Police Station',
    'fir_no': 'FIR No.', 'efir_no': 'E-FIR No.',
    'dd_nofir_no': 'DD No./FIR No.', 'firdd_no': 'FIR/DD No.',
    'dd_no': 'DD No.', 'dd_date': 'DD Date',
    'us': 'U/S',
    'io': 'Name of IO', 'name_of_io': 'Name of IO', 'io_name': 'IO Name',
    'io_mobile_no': 'IO Mobile No.',
    'rank_of_io': 'Rank of IO', 'mobile_no_of_io': 'Mobile No. of IO',
    'complainant_details': 'Complainant (Name / S/O / Address)',
    'arrested_details': 'Arrested Person (Name / Age / S/O / Address)',
    'accused_details': 'Accused (Name / Age / S/O / Address)',
    'po_details': 'PO Details (Name / S/O / Address)',
    'deceased_details': 'Deceased (Name / Age / S/O / Address)',
    'traced_person_details': 'Traced Person (Name / S/O / Address)',
    'place_of_occurrence': 'Place of Occurrence',
    'place_of_occurrence_1': 'Place of Occurrence (2)',
    'time_of_occurrence': 'Time of Occurrence',
    'date_of_occurrence': 'Date of Occurrence',
    'pcjcbail': 'PC/JC/Bail',
    'recovery': 'Recovery',
    'beat_no': 'Beat No.',
    'stolen_items': 'Stolen Items',
    'cause_of_death': 'Cause of Death',
    'challan_untrace_cancel': 'Challan / Untrace / Cancel',
    'name_of_operator_to_whom_mps': 'Name of Operator (MPS)',
    'found_place': 'Found Place', 'found_date': 'Found Date',
    'upper_dress_color': 'Upper Dress Color',
    'lower_dress_color': 'Lower Dress Color',
    'prev_involvement_no_of_cases': 'Prev. Involvement (No. of Cases)',
    'whether_accused_is_bc_or_not': 'Accused BC?',
    'group_patrolling': 'Group Patrolling',
    'cycle_patrolling': 'Cycle Patrolling',
    'by_antisnatching_team': 'By Anti-Snatching Team',
    'by_prahari': 'By Prahari',
    'by_eyes_ears_scheme_members': 'By Eyes & Ears Scheme Members',
    'pcr_call': 'PCR Call',
}


def build_workbook(sheets_data, sheets, file_path, date=''):
    """
    Build a daily-diary Excel workbook and write it to file_path.

    sheets_data: { table_name: [row_dict, ...] }
    sheets:      list of sheet descriptor dicts from registry.SHEETS
    file_path:   absolute path for the output .xlsx file
    date:        optional date string shown in sheet headers
    """
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    for sheet_def in sheets:
        table_name   = sheet_def['table_name']
        label        = sheet_def['label']
        num          = sheet_def['num']
        col_keys     = sheet_def['columns']
        col_labels   = sheet_def.get('column_labels', {})

        rows = sheets_data.get(table_name, [])

        sheet_title = f'{num}. {label}'
        for bad_char in ':\\/?*[]':
            sheet_title = sheet_title.replace(bad_char, '')
        sheet_title = sheet_title[:31]

        ws = wb.create_sheet(title=sheet_title)

        # Title and date rows
        ws.append([f'PHAROS Daily Diary — {label}'])
        ws.cell(row=ws.max_row, column=1).font = _TITLE_FONT
        ws.append([f'Date: {date}'] if date else [''])
        ws.cell(row=ws.max_row, column=1).font = _META_FONT
        ws.append([])  # spacer

        # Header row
        col_headers = [_resolve_header(k, col_labels, SHARED_COLUMN_LABELS) for k in col_keys]
        ws.append(col_headers if col_headers else ['No Data'])
        header_row_idx = ws.max_row
        ws.row_dimensions[header_row_idx].height = 22
        for ci in range(1, len(col_headers) + 1):
            cell = ws.cell(row=header_row_idx, column=ci)
            cell.fill   = _HEADER_FILL
            cell.font   = _HEADER_FONT
            cell.alignment = _HEADER_ALIGN
            cell.border = _THIN_BORDER

        # Data rows
        if not rows:
            ws.append(['No records for this date.'])
        else:
            for row_offset, row_dict in enumerate(rows, start=1):
                row_values = [row_dict.get(k, '') for k in col_keys]
                ws.append(row_values)
                actual_row = header_row_idx + row_offset
                ws.row_dimensions[actual_row].height = 18
                use_zebra = (row_offset % 2 == 0)
                for ci, val in enumerate(row_values, start=1):
                    cell = ws.cell(row=actual_row, column=ci)
                    cell.font      = _DATA_FONT
                    cell.alignment = _DATA_ALIGN
                    cell.border    = _THIN_BORDER
                    if use_zebra:
                        cell.fill = _ZEBRA_FILL

        # Auto-fit column widths (cap at 45)
        for col_cells in ws.columns:
            max_len = 10
            for cell in col_cells:
                if cell.row < header_row_idx:
                    continue
                val_str = str(cell.value or '')
                if len(val_str) > max_len:
                    max_len = len(val_str)
            col_letter = openpyxl.utils.get_column_letter(col_cells[0].column)
            ws.column_dimensions[col_letter].width = min(max_len + 3, 45)

    if not wb.worksheets:
        ws = wb.create_sheet(title='No Data')
        ws.append(['No records found for the selected date and filters.'])

    wb.save(file_path)
    print(f'[Builder] Workbook saved: {file_path} ({len(wb.worksheets)} sheets)')
