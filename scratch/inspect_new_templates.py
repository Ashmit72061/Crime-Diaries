import openpyxl
import os

def inspect_template(filename):
    print(f"=== Inspecting {filename} ===")
    if not os.path.exists(filename):
        print("File does not exist")
        return
    wb = openpyxl.load_workbook(filename)
    sheet = wb.active
    print(f"Sheet Name: {sheet.title}")
    print(f"Max Row: {sheet.max_row}, Max Col: {sheet.max_column}")
    for r in range(1, min(6, sheet.max_row + 1)):
        row_vals = [cell.value for cell in sheet[r]]
        print(f"Row {r}: {row_vals[:30]}")
    print("-" * 50)

inspect_template("CASE_Import_Template (4).xlsx")
inspect_template("ARREST_Import_Template (2).xlsx")
